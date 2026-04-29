"""
GST report generation service for IRAS filing.

Singapore GST Formulas:
  - Standard: GST = price * 9 / 109, ROUND_HALF_UP
  - Zero-rated: Thomson entity = 0 GST

IRAS Formats:
  - GST9: Standard filing (most businesses)
  - GST109: Simplified filing (smaller businesses)

Compliance Requirements:
  - Zero AI interpolation
  - Deterministic calculations
  - Immutable reports once generated
  - Thomson entity: 0% GST exempt
"""
import io
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from typing import TYPE_CHECKING, List
from uuid import UUID

from django.db.models import Sum

if TYPE_CHECKING:
    from ..models import Transaction


# Singapore GST rate: 9%
GST_RATE = Decimal("9")
GST_BASE = Decimal("109")  # 100 + 9


@dataclass(frozen=True)
class GSTTransaction:
    """GST transaction detail for reporting."""
    description: str
    value: Decimal
    gst_amount: Decimal
    date: date
    source: str  # Agreement ID or "Manual"


@dataclass(frozen=True)
class GSTReportResult:
    """GST report calculation result."""
    entity_id: UUID
    quarter: str
    total_sales: Decimal
    total_gst: Decimal
    transactions: List[GSTTransaction]
    format_type: str  # "GST9" or "GST109"


def extract_gst(amount: Decimal, entity_code: str = "") -> Decimal:
    """
    Extract GST from amount using Singapore formula.

    Formula: GST = price * 9 / 109, ROUND_HALF_UP

    Args:
        amount: Total amount including GST
        entity_code: Entity code (for Thomson exemption check)

    Returns:
        GST amount (0.00 for Thomson entity)

    Examples:
        >>> extract_gst(Decimal("109.00"))
        Decimal('9.00')
        >>> extract_gst(Decimal("218.00"))
        Decimal('18.00')
        >>> extract_gst(Decimal("50.00"))
        Decimal('4.13')
        >>> extract_gst(Decimal("100.00"), "THOMSON")
        Decimal('0.00')
    """
    # Thomson entity is GST exempt
    if entity_code.upper() == "THOMSON":
        return Decimal("0.00")

    # GST = price * 9 / 109, ROUND_HALF_UP
    gst = (amount * GST_RATE / GST_BASE).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    return gst


def extract_gst_components(
    entity_id: UUID,
    quarter: str,
) -> List[GSTTransaction]:
    """
    Extract GST components from SalesAgreement for period.

    Args:
        entity_id: Entity UUID
        quarter: Quarter string (e.g., "2026-Q1")

    Returns:
        List of GST transactions
    """
    from apps.sales.models import SalesAgreement, AgreementStatus
    from apps.core.models import Entity

    # Parse quarter
    year, q = quarter.split("-Q")
    year = int(year)
    quarter_num = int(q)

    # Determine quarter date range
    quarter_months = {
        1: (1, 3),
        2: (4, 6),
        3: (7, 9),
        4: (10, 12),
    }
    start_month, end_month = quarter_months[quarter_num]
    start_date = date(year, start_month, 1)
    end_date = date(year, end_month, 31 if end_month in [3, 12] else 30)

    # Get entity code for Thomson check
    try:
        entity = Entity.objects.get(id=entity_id)
        entity_code = entity.code
    except Entity.DoesNotExist:
        entity_code = ""

    # Get completed agreements in quarter
    agreements = SalesAgreement.objects.filter(
        entity_id=entity_id,
        status=AgreementStatus.COMPLETED,
        completed_at__date__gte=start_date,
        completed_at__date__lte=end_date,
    ).select_related("buyer")

    transactions = []
    for agreement in agreements:
        total = agreement.total_amount or Decimal("0.00")
        gst = extract_gst(total, entity_code)

        transactions.append(
            GSTTransaction(
                description=f"Sale to {agreement.buyer.name}",
                value=total - gst,  # Value excluding GST
                gst_amount=gst,
                date=agreement.completed_at.date() if agreement.completed_at else start_date,
                source=str(agreement.id),
            )
        )

    # Also include manual transactions with GST
    from ..models import Transaction, TransactionType

    manual_transactions = Transaction.objects.filter(
        entity_id=entity_id,
        type=TransactionType.REVENUE,
        date__gte=start_date,
        date__lte=end_date,
        gst_component__gt=Decimal("0.00"),
    )

    for txn in manual_transactions:
        transactions.append(
            GSTTransaction(
                description=txn.description or "Manual transaction",
                value=txn.amount - txn.gst_component,
                gst_amount=txn.gst_component,
                date=txn.date,
                source="Manual",
            )
        )

    return transactions


def gen_gst_report(
    entity_id: UUID,
    quarter: str,
    format_type: str = "GST9",
) -> GSTReportResult:
    """
    Generate GST report for IRAS filing.

    Args:
        entity_id: Entity UUID
        quarter: Quarter string (e.g., "2026-Q1")
        format_type: "GST9" or "GST109"

    Returns:
        GSTReportResult with report data
    """
    transactions = extract_gst_components(entity_id, quarter)

    total_sales = sum(t.value for t in transactions)
    total_gst = sum(t.gst_amount for t in transactions)

    return GSTReportResult(
        entity_id=entity_id,
        quarter=quarter,
        total_sales=total_sales,
        total_gst=total_gst,
        transactions=transactions,
        format_type=format_type,
    )


def gen_gst_excel(
    entity_id: UUID,
    quarter: str,
    format_type: str = "GST9",
) -> bytes:
    """
    Generate GST report as Excel file.

    Args:
        entity_id: Entity UUID
        quarter: Quarter string
        format_type: "GST9" or "GST109"

    Returns:
        Excel file as bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from apps.core.models import Entity

    report = gen_gst_report(entity_id, quarter, format_type)

    wb = Workbook()
    ws = wb.active
    ws.title = f"GST {quarter}"

    # Header
    entity = Entity.objects.get(id=entity_id)
    ws["A1"] = f"GST Report - {quarter}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Entity: {entity.name}"
    ws["A3"] = f"Format: {format_type}"

    # Column headers
    headers = ["Date", "Description", "Value (Excl. GST)", "GST Amount", "Source"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    # Data rows
    row = 6
    for txn in report.transactions:
        ws.cell(row=row, column=1, value=txn.date.isoformat())
        ws.cell(row=row, column=2, value=txn.description)
        ws.cell(row=row, column=3, value=float(txn.value))
        ws.cell(row=row, column=4, value=float(txn.gst_amount))
        ws.cell(row=row, column=5, value=txn.source)
        row += 1

    # Summary
    ws.cell(row=row + 1, column=2, value="Total Sales (Excl. GST):")
    ws.cell(row=row + 1, column=3, value=float(report.total_sales))
    ws.cell(row=row + 2, column=2, value="Total GST Amount:")
    ws.cell(row=row + 2, column=4, value=float(report.total_gst))
    ws.cell(row=row + 2, column=4).font = Font(bold=True)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def gen_pnl_excel(entity_id: UUID, month: date) -> bytes:
    """
    Generate P&L statement as Excel file.

    Args:
        entity_id: Entity UUID
        month: First day of month

    Returns:
        Excel file as bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from apps.core.models import Entity
    from .pnl import calc_pnl

    result = calc_pnl(entity_id, month)
    entity = Entity.objects.get(id=entity_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"P&L {month.strftime('%Y-%m')}"

    # Header
    ws["A1"] = f"Profit & Loss Statement"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Entity: {entity.name}"
    ws["A3"] = f"Month: {month.strftime('%B %Y')}"

    # P&L Items
    data = [
        ("Revenue", result.revenue),
        ("", ""),
        ("Less: Cost of Goods Sold", result.cogs),
        ("Less: Operating Expenses", result.expenses),
        ("", ""),
        ("Net Profit/(Loss)", result.net),
        ("", ""),
        ("YTD Revenue", result.ytd_revenue),
        ("YTD Net Profit/(Loss)", result.ytd_net),
    ]

    for row, (label, value) in enumerate(data, 5):
        ws.cell(row=row, column=1, value=label)
        if isinstance(value, Decimal):
            ws.cell(row=row, column=2, value=float(value))
        else:
            ws.cell(row=row, column=2, value=value)

        # Bold for totals
        if label in ["Revenue", "Net Profit/(Loss)", "YTD Revenue", "YTD Net Profit/(Loss)"]:
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)

    # Category breakdown
    if result.by_category:
        ws.cell(row=row + 2, column=1, value="Expense by Category:").font = Font(bold=True)
        row += 3
        for category, amount in result.by_category.items():
            ws.cell(row=row, column=1, value=f"  {category}")
            ws.cell(row=row, column=2, value=float(amount))
            row += 1

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def validate_gst_calculation(price: Decimal, expected_gst: Decimal) -> bool:
    """
    Validate GST calculation matches expected value.

    Args:
        price: Total price including GST
        expected_gst: Expected GST amount

    Returns:
        True if calculation matches

    Examples:
        >>> validate_gst_calculation(Decimal("109.00"), Decimal("9.00"))
        True
        >>> validate_gst_calculation(Decimal("218.00"), Decimal("18.00"))
        True
        >>> validate_gst_calculation(Decimal("50.00"), Decimal("4.13"))
        True
    """
    calculated = extract_gst(price)
    return calculated == expected_gst
