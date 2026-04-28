"""GST Router
=============
Phase 6: Compliance & NParks Reporting

Django Ninja router for GST reporting endpoints.
"""

import logging
from datetime import date
from decimal import Decimal
from typing import Optional
from uuid import UUID

from django.http import HttpResponse
from ninja import Router
from ninja.errors import HttpError

from apps.core.auth import AuthenticationService
from apps.core.permissions import require_role

from ..models import GSTLedger
from ..schemas import (
    GSTCalculationRequest,
    GSTCalculationResponse,
    GSTSummary,
    GSTLedgerResponse,
    GSTLedgerEntry,
)
from ..services.gst import GSTService

logger = logging.getLogger(__name__)

router = Router(tags=["compliance", "gst"])


@router.post("/calculate", response=GSTCalculationResponse)
def calculate_gst(
    request,
    data: GSTCalculationRequest,
):
    """
    Calculate GST for a given price.
    
    Uses Singapore GST formula: price * 9 / 109, ROUND_HALF_UP.
    Thomson entity returns 0% GST (exempt).
    """
    user = AuthenticationService.get_user_from_request(request)
    if not user:
        raise HttpError(401, "Authentication required")
    
    try:
        result = GSTService.calculate_gst_from_request(data)
        return result
    except Exception as e:
        logger.error(f"GST calculation failed: {e}")
        raise HttpError(400, str(e))


@router.get("/summary", response=GSTSummary)
def get_gst_summary(
    request,
    entity_id: UUID,
    quarter: str,  # Format: YYYY-Q#
):
    """
    Get GST summary for entity and quarter.
    """
    user = AuthenticationService.get_user_from_request(request)
    if not user:
        raise HttpError(401, "Authentication required")
    
    # Validate quarter format
    try:
        year, q = quarter.split("-")
        year = int(year)
        q_num = int(q.replace("Q", ""))
        if q_num < 1 or q_num > 4:
            raise ValueError("Invalid quarter number")
    except (ValueError, IndexError):
        raise HttpError(400, "Invalid quarter format. Use YYYY-Q# (e.g., 2026-Q1)")
    
    # Check access
    from apps.core.models import Entity
    
    try:
        entity = Entity.objects.get(id=entity_id)
    except Entity.DoesNotExist:
        raise HttpError(404, "Entity not found")
    
    if not user.has_role("management") and user.entity_id != entity_id:
        raise HttpError(403, "Access denied")
    
    try:
        summary = GSTService.calc_gst_summary(entity, quarter)
        return summary
    except Exception as e:
        logger.error(f"GST summary failed: {e}")
        raise HttpError(400, str(e))


@router.get("/ledger", response=GSTLedgerResponse)
def get_gst_ledger(
    request,
    entity_id: UUID,
    quarter: str,
    page: int = 1,
    per_page: int = 25,
):
    """
    Get GST ledger entries for entity and quarter.
    """
    user = AuthenticationService.get_user_from_request(request)
    if not user:
        raise HttpError(401, "Authentication required")
    
    # Validate quarter format
    try:
        year, q = quarter.split("-")
        year = int(year)
        q_num = int(q.replace("Q", ""))
        if q_num < 1 or q_num > 4:
            raise ValueError("Invalid quarter number")
    except (ValueError, IndexError):
        raise HttpError(400, "Invalid quarter format. Use YYYY-Q# (e.g., 2026-Q1)")
    
    # Check access
    from apps.core.models import Entity
    
    try:
        entity = Entity.objects.get(id=entity_id)
    except Entity.DoesNotExist:
        raise HttpError(404, "Entity not found")
    
    if not user.has_role("management") and user.entity_id != entity_id:
        raise HttpError(403, "Access denied")
    
    try:
        entries = GSTService.get_ledger_entries(entity, quarter)
        
        # Pagination
        total = len(entries)
        start = (page - 1) * per_page
        end = start + per_page
        paginated = entries[start:end]
        
        return GSTLedgerResponse(
            items=paginated,
            total=total,
            page=page,
            per_page=per_page,
        )
    except Exception as e:
        logger.error(f"GST ledger failed: {e}")
        raise HttpError(400, str(e))


@router.get("/export")
def export_gst_ledger(
    request,
    entity_id: UUID,
    quarter: str,
):
    """
    Export GST ledger to Excel.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    user = AuthenticationService.get_user_from_request(request)
    if not user:
        raise HttpError(401, "Authentication required")
    
    # Validate quarter format
    try:
        year, q = quarter.split("-")
        year = int(year)
        q_num = int(q.replace("Q", ""))
        if q_num < 1 or q_num > 4:
            raise ValueError("Invalid quarter number")
    except (ValueError, IndexError):
        raise HttpError(400, "Invalid quarter format. Use YYYY-Q# (e.g., 2026-Q1)")
    
    # Check access
    from apps.core.models import Entity
    
    try:
        entity = Entity.objects.get(id=entity_id)
    except Entity.DoesNotExist:
        raise HttpError(404, "Entity not found")
    
    if not user.has_role("management") and user.entity_id != entity_id:
        raise HttpError(403, "Access denied")
    
    try:
        # Get entries
        entries = GSTService.get_ledger_entries(entity, quarter)
        summary = GSTService.calc_gst_summary(entity, quarter)
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"GST {quarter}"
        
        # Header
        ws.cell(row=1, column=1, value=f"GST Ledger - {entity.name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=2, column=1, value=f"Period: {quarter}")
        ws.cell(row=2, column=1).font = Font(bold=True)
        
        ws.cell(row=3, column=1, value=f"Total Sales: ${summary.total_sales}")
        ws.cell(row=4, column=1, value=f"Total GST: ${summary.total_gst}")
        ws.cell(row=5, column=1, value=f"Transactions: {summary.transactions_count}")
        
        # Column headers
        headers = ["Agreement #", "Buyer", "Total Sales", "GST Component", "Date"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row_num, entry in enumerate(entries, 8):
            ws.cell(row=row_num, column=1, value=entry.agreement_number)
            ws.cell(row=row_num, column=2, value=entry.buyer_name)
            ws.cell(row=row_num, column=3, value=float(entry.total_sales))
            ws.cell(row=row_num, column=4, value=float(entry.gst_component))
            ws.cell(row=row_num, column=5, value=entry.created_at.strftime("%Y-%m-%d"))
        
        # Auto-adjust columns
        from openpyxl.utils import get_column_letter
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"gst_ledger_{entity.code.lower()}_{quarter}.xlsx"
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        logger.error(f"GST export failed: {e}")
        raise HttpError(400, str(e))


@router.get("/quarters")
def get_available_quarters(
    request,
    entity_id: Optional[UUID] = None,
):
    """
    Get available GST quarters for entity.
    Returns list of quarters with data.
    """
    user = AuthenticationService.get_user_from_request(request)
    if not user:
        raise HttpError(401, "Authentication required")
    
    # Check access
    from apps.core.models import Entity
    
    if entity_id:
        try:
            entity = Entity.objects.get(id=entity_id)
        except Entity.DoesNotExist:
            raise HttpError(404, "Entity not found")
        
        if not user.has_role("management") and user.entity_id != entity_id:
            raise HttpError(403, "Access denied")
    else:
        if not user.entity_id:
            raise HttpError(400, "Entity ID required")
        entity_id = user.entity_id
    
    # Get distinct quarters from ledger
    quarters = GSTLedger.objects.filter(
        entity_id=entity_id,
    ).values_list("period", flat=True).distinct().order_by("-period")
    
    return {
        "quarters": list(quarters),
        "entity_id": str(entity_id),
    }
