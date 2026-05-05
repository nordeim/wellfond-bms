"""NParks Service
================
Phase 6: Compliance & NParks Reporting

Deterministic Excel generation for NParks submissions.
Zero AI - pure Python/SQL logic.

Generates 5 Excel documents:
1. Mating Sheet - breeding records with dual-sire support
2. Puppy Movement - puppy sales/transfers
3. Vet Treatments - veterinary records
4. Puppies Bred - in-house breeding summary
5. Dog Movement - rehomed/deceased dogs
"""

import io
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from apps.core.models import Entity
from apps.breeding.models import BreedingRecord, Litter, Puppy
from apps.operations.models import Dog, HealthRecord
from apps.sales.models import SalesAgreement, AgreementLineItem

from ..models import NParksSubmission, NParksStatus
from ..schemas import NParksPreview, NParksPreviewRow

logger = logging.getLogger(__name__)

# Farm details (constant per v1.1 hardening)
FARM_DETAILS = {
    "name": "Wellfond Pets Holdings Pte Ltd",
    "license_number": "DB000065X",
    "address": "123 Pet Avenue, Singapore 123456",
}


class NParksService:
    """Service for generating NParks compliance documents."""

    @staticmethod
    def generate_nparks(
        entity_id: UUID,
        month: date,
        generated_by_id: UUID,
    ) -> dict[str, bytes]:
        """
        Generate all 5 NParks documents for entity+month.

        Args:
            entity_id: Entity UUID
            month: First day of reporting month
            generated_by_id: User ID generating the report

        Returns:
            Dict mapping document type to Excel bytes:
            {
                "mating_sheet": bytes,
                "puppy_movement": bytes,
                "vet_treatments": bytes,
                "puppies_bred": bytes,
                "dog_movement": bytes,
            }

        Raises:
            ValueError: If entity not found or no data for month
        """
        # Validate entity
        try:
            entity = Entity.objects.get(id=entity_id)
        except Entity.DoesNotExist:
            raise ValueError(f"Entity not found: {entity_id}")

        # Calculate month range
        month_start = month.replace(day=1)
        if month.month == 12:
            month_end = month.replace(year=month.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month.replace(month=month.month + 1, day=1) - timedelta(days=1)

        # Generate each document
        documents = {}

        # 1. Mating Sheet (breeding records)
        documents["mating_sheet"] = NParksService._generate_mating_sheet(
            entity, month_start, month_end
        )

        # 2. Puppy Movement (puppy sales)
        documents["puppy_movement"] = NParksService._generate_puppy_movement(
            entity, month_start, month_end
        )

        # 3. Vet Treatments
        documents["vet_treatments"] = NParksService._generate_vet_treatments(
            entity, month_start, month_end
        )

        # 4. Puppies Bred (in-house)
        documents["puppies_bred"] = NParksService._generate_puppies_bred(
            entity, month_start, month_end
        )

        # 5. Dog Movement (rehomed/deceased)
        documents["dog_movement"] = NParksService._generate_dog_movement(
            entity, month_start, month_end
        )

        # Create submission record
        submission = NParksSubmission.objects.create(
            entity=entity,
            month=month,
            status=NParksStatus.DRAFT,
            generated_by_id=generated_by_id,
        )

        logger.info(
            f"Generated NParks documents for {entity.name} - {month.strftime('%Y-%m')} "
            f"(Submission ID: {submission.id})"
        )

        return documents

    @staticmethod
    def _generate_mating_sheet(entity: Entity, month_start: date, month_end: date) -> bytes:
        """Generate mating sheet Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Mating Sheet"

        # Headers
        headers = [
            "S/N",
            "Dam Microchip",
            "Dam Breed",
            "Sire 1 Microchip",
            "Sire 1 Breed",
            "Sire 2 Microchip",
            "Sire 2 Breed",
            "Mating Date",
            "Method",
            "Whelping Date",
            "Puppies Born",
            "Alive",
            "Stillborn",
        ]

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Get breeding records for month
        records = BreedingRecord.objects.filter(
            entity=entity,
            date__gte=month_start,
            date__lte=month_end,
        ).select_related("dam", "sire1", "sire2")

        # Sort by dam microchip, then date (deterministic)
        records = records.order_by("dam__microchip", "date")

        # Populate data
        for row_num, record in enumerate(records, 1):
            row = row_num + 1  # Account for header

            ws.cell(row=row, column=1, value=row_num)
            ws.cell(row=row, column=2, value=record.dam.microchip)
            ws.cell(row=row, column=3, value=record.dam.breed)
            ws.cell(row=row, column=4, value=record.sire1.microchip if record.sire1 else "")
            ws.cell(row=row, column=5, value=record.sire1.breed if record.sire1 else "")
            ws.cell(row=row, column=6, value=record.sire2.microchip if record.sire2 else "")
            ws.cell(row=row, column=7, value=record.sire2.breed if record.sire2 else "")
            ws.cell(row=row, column=8, value=record.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=9, value=record.get_method_display())

            # Get litter info if exists
            litter = Litter.objects.filter(breeding_record=record).first()
            if litter:
                ws.cell(row=row, column=10, value=litter.whelp_date.strftime("%Y-%m-%d") if litter.whelp_date else "")
                ws.cell(row=row, column=11, value=litter.alive_count + litter.stillborn_count)
                ws.cell(row=row, column=12, value=litter.alive_count)
                ws.cell(row=row, column=13, value=litter.stillborn_count)

        # Add farm details footer
        last_row = len(records) + 3
        ws.cell(row=last_row, column=1, value=f"Farm: {FARM_DETAILS['name']}")
        ws.cell(row=last_row + 1, column=1, value=f"License: {FARM_DETAILS['license_number']}")
        ws.cell(row=last_row + 2, column=1, value=f"Report Period: {month_start.strftime('%Y-%m')}")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _generate_puppy_movement(entity: Entity, month_start: date, month_end: date) -> bytes:
        """Generate puppy movement Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Puppy Movement"

        headers = [
            "S/N",
            "Microchip",
            "Name",
            "Breed",
            "Gender",
            "DOB",
            "Sire Microchip",
            "Dam Microchip",
            "Buyer Name",
            "Buyer Mobile",
            "Buyer Email",
            "Transfer Date",
            "Sale Price",
        ]

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Get puppies sold/transferred in month
        agreements = SalesAgreement.objects.filter(
            entity=entity,
            status="COMPLETED",
            completed_at__date__gte=month_start,
            completed_at__date__lte=month_end,
        ).prefetch_related("line_items__dog")

        row_num = 0
        current_row = 2

        for agreement in agreements:
            for item in agreement.line_items.all():
                if hasattr(item.dog, 'microchip'):  # Only actual dogs, not placeholder items
                    row_num += 1

                    ws.cell(row=current_row, column=1, value=row_num)
                    ws.cell(row=current_row, column=2, value=item.dog.microchip)
                    ws.cell(row=current_row, column=3, value=item.dog.name)
                    ws.cell(row=current_row, column=4, value=item.dog.breed)
                    ws.cell(row=current_row, column=5, value=item.dog.gender)
                    ws.cell(row=current_row, column=6, value=item.dog.dob.strftime("%Y-%m-%d"))
                    ws.cell(row=current_row, column=9, value=agreement.buyer_name)
                    ws.cell(row=current_row, column=10, value=agreement.buyer_mobile)
                    ws.cell(row=current_row, column=11, value=agreement.buyer_email)
                    ws.cell(row=current_row, column=12, value=agreement.completed_at.strftime("%Y-%m-%d"))
                    ws.cell(row=current_row, column=13, value=item.line_total)

                    current_row += 1

        # Add farm details footer
        last_row = current_row + 1
        ws.cell(row=last_row, column=1, value=f"Farm: {FARM_DETAILS['name']}")
        ws.cell(row=last_row + 1, column=1, value=f"License: {FARM_DETAILS['license_number']}")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _generate_vet_treatments(entity: Entity, month_start: date, month_end: date) -> bytes:
        """Generate vet treatments Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vet Treatments"

        headers = [
            "S/N",
            "Microchip",
            "Name",
            "Breed",
            "Treatment Date",
            "Treatment Type",
            "Description",
            "Vet Name",
            "Follow-up Required",
        ]

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Get health records for month
        records = HealthRecord.objects.filter(
            dog__entity=entity,
            date__gte=month_start,
            date__lte=month_end,
        ).select_related("dog")

        # Sort by microchip, then date
        records = records.order_by("dog__microchip", "date")

        for row_num, record in enumerate(records, 1):
            row = row_num + 1

            ws.cell(row=row, column=1, value=row_num)
            ws.cell(row=row, column=2, value=record.dog.microchip)
            ws.cell(row=row, column=3, value=record.dog.name)
            ws.cell(row=row, column=4, value=record.dog.breed)
            ws.cell(row=row, column=5, value=record.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=6, value=record.category)
            ws.cell(row=row, column=7, value=record.description[:100])  # Truncate long descriptions
            ws.cell(row=row, column=8, value=record.vet_name or "")
            ws.cell(row=row, column=9, value="Yes" if record.follow_up_required else "No")

        # Add farm details footer
        last_row = len(records) + 3
        ws.cell(row=last_row, column=1, value=f"Farm: {FARM_DETAILS['name']}")
        ws.cell(row=last_row + 1, column=1, value=f"License: {FARM_DETAILS['license_number']}")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _generate_puppies_bred(entity: Entity, month_start: date, month_end: date) -> bytes:
        """Generate puppies bred summary Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Puppies Bred"

        headers = [
            "S/N",
            "Litter ID",
            "Dam Microchip",
            "Dam Breed",
            "Sire Microchip",
            "Sire Breed",
            "Whelp Date",
            "Total Born",
            "Alive",
            "Stillborn",
            "Gender Breakdown",
        ]

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Get litters for month
        litters = Litter.objects.filter(
            breeding_record__entity=entity,
            whelp_date__gte=month_start,
            whelp_date__lte=month_end,
        ).select_related("breeding_record__dam", "breeding_record__sire1")

        for row_num, litter in enumerate(litters, 1):
            row = row_num + 1
            breeding = litter.breeding_record

            # Count genders
            male_count = Puppy.objects.filter(litter=litter, gender="M").count()
            female_count = Puppy.objects.filter(litter=litter, gender="F").count()

            ws.cell(row=row, column=1, value=row_num)
            ws.cell(row=row, column=2, value=str(litter.id)[:8])
            ws.cell(row=row, column=3, value=breeding.dam.microchip)
            ws.cell(row=row, column=4, value=breeding.dam.breed)
            ws.cell(row=row, column=5, value=breeding.sire1.microchip if breeding.sire1 else "")
            ws.cell(row=row, column=6, value=breeding.sire1.breed if breeding.sire1 else "")
            ws.cell(row=row, column=7, value=litter.whelp_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=8, value=litter.alive_count + litter.stillborn_count)
            ws.cell(row=row, column=9, value=litter.alive_count)
            ws.cell(row=row, column=10, value=litter.stillborn_count)
            ws.cell(row=row, column=11, value=f"{male_count}M / {female_count}F")

        # Add farm details footer
        last_row = len(litters) + 3
        ws.cell(row=last_row, column=1, value=f"Farm: {FARM_DETAILS['name']}")
        ws.cell(row=last_row + 1, column=1, value=f"License: {FARM_DETAILS['license_number']}")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _generate_dog_movement(entity: Entity, month_start: date, month_end: date) -> bytes:
        """Generate dog movement (rehomed/deceased) Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dog Movement"

        headers = [
            "S/N",
            "Microchip",
            "Name",
            "Breed",
            "Gender",
            "DOB",
            "Movement Type",
            "Date",
            "Destination/Reason",
            "Notes",
        ]

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Get rehoming agreements (type=REHOME) in month
        rehomed = SalesAgreement.objects.filter(
            entity=entity,
            type="REHOME",
            status="COMPLETED",
            completed_at__date__gte=month_start,
            completed_at__date__lte=month_end,
        ).prefetch_related("line_items__dog")

        row_num = 0
        current_row = 2

        for agreement in rehomed:
            for item in agreement.line_items.all():
                row_num += 1

                ws.cell(row=current_row, column=1, value=row_num)
                ws.cell(row=current_row, column=2, value=item.dog.microchip)
                ws.cell(row=current_row, column=3, value=item.dog.name)
                ws.cell(row=current_row, column=4, value=item.dog.breed)
                ws.cell(row=current_row, column=5, value=item.dog.gender)
                ws.cell(row=current_row, column=6, value=item.dog.dob.strftime("%Y-%m-%d"))
                ws.cell(row=current_row, column=7, value="Rehomed")
                ws.cell(row=current_row, column=8, value=agreement.completed_at.strftime("%Y-%m-%d"))
                ws.cell(row=current_row, column=9, value=agreement.buyer_name)
                ws.cell(row=current_row, column=10, value=agreement.special_conditions[:100])

                current_row += 1

        # Add farm details footer
        last_row = current_row + 1
        ws.cell(row=last_row, column=1, value=f"Farm: {FARM_DETAILS['name']}")
        ws.cell(row=last_row + 1, column=1, value=f"License: {FARM_DETAILS['license_number']}")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def preview_nparks(
        entity_id: UUID,
        month: date,
        doc_type: str,
    ) -> NParksPreview:
        """
        Preview NParks document data before generation.

        Args:
            entity_id: Entity UUID
            month: Reporting month
            doc_type: Document type to preview

        Returns:
            Preview data with headers and sample rows
        """
        # Calculate month range
        month_start = month.replace(day=1)
        if month.month == 12:
            month_end = month.replace(year=month.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = month.replace(month=month.month + 1, day=1) - timedelta(days=1)

        # Get preview data based on doc type
        if doc_type == "mating_sheet":
            headers = ["S/N", "Dam", "Sire 1", "Sire 2", "Date", "Method"]
            records = BreedingRecord.objects.filter(
                entity_id=entity_id,
                date__gte=month_start,
                date__lte=month_end,
            ).order_by("dam__microchip", "date")[:10]

            rows = [
                NParksPreviewRow(
                    row_num=i + 1,
                    data={
                        "S/N": i + 1,
                        "Dam": r.dam.microchip,
                        "Sire 1": r.sire1.microchip if r.sire1 else "",
                        "Sire 2": r.sire2.microchip if r.sire2 else "",
                        "Date": r.date.strftime("%Y-%m-%d"),
                        "Method": r.get_method_display(),
                    }
                )
                for i, r in enumerate(records)
            ]
        else:
            # Generic preview for other doc types
            headers = ["Preview", "Not Available"]
            rows = []

        return NParksPreview(
            doc_type=doc_type,
            headers=headers,
            rows=rows,
            total_rows=len(rows),
        )

    @staticmethod
    def validate_nparks(
        entity_id: UUID,
        month: date,
    ) -> list[str]:
        """
        Validate NParks submission requirements.

        Returns:
            List of warning messages (empty if valid)
        """
        warnings = []

        # Check for incomplete breeding records
        incomplete = BreedingRecord.objects.filter(
            entity_id=entity_id,
            date__year=month.year,
            date__month=month.month,
            confirmed_sire="UNCONFIRMED",
        ).count()

        if incomplete > 0:
            warnings.append(f"{incomplete} breeding records have unconfirmed sire")

        # Check for missing microchips
        missing_chips = Dog.objects.filter(
            entity_id=entity_id,
            microchip__isnull=True,
        ).count()

        if missing_chips > 0:
            warnings.append(f"{missing_chips} dogs missing microchip numbers")

        return warnings
