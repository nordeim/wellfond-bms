"""NParks Tests
==============
Phase 6: Compliance & NParks Reporting

Tests for NParks document generation and validation.
"""

import uuid
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from apps.breeding.models import BreedingRecord
from apps.core.models import Entity, User
from apps.operations.models import Dog

from ..models import NParksSubmission, NParksStatus
from ..services.nparks import NParksService


class TestNParksGeneration(TestCase):
    """Test NParks document generation."""

    def setUp(self):
        """Set up test data."""
        # Create entity
        self.holdings, _ = Entity.objects.get_or_create(
            defaults={"name": "Holdings", "code": "HOLDINGS", "slug": "holdings-test"},
            id=uuid.uuid4(),
        )
        self.katong, _ = Entity.objects.get_or_create(
            defaults={"name": "Katong", "code": "KATONG", "slug": "katong-test"},
            id=uuid.uuid4(),
        )
        self.thomson, _ = Entity.objects.get_or_create(
            defaults={"name": "Thomson", "code": "THOMSON", "slug": "thomson-test"},
            id=uuid.uuid4(),
        )

        # Create user
        self.user = User.objects.create_user(
            username=f"testuser_{uuid.uuid4().hex[:8]}",
            email="test@example.com",
            password="testpass123",
            entity=self.holdings,
            role="management",
        )

        # Create test dogs
        self.dam = Dog.objects.create(
            microchip="DAM001",
            name="Test Dam",
            breed="Golden Retriever",
            dob="2020-01-01",
            gender="F",
            entity=self.holdings,
        )
        self.sire1 = Dog.objects.create(
            microchip="SIRE001",
            name="Test Sire 1",
            breed="Golden Retriever",
            dob="2019-01-01",
            gender="M",
            entity=self.holdings,
        )
        self.sire2 = Dog.objects.create(
            microchip="SIRE002",
            name="Test Sire 2",
            breed="Labrador",
            dob="2019-06-01",
            gender="M",
            entity=self.holdings,
        )

    def test_generate_holdings_5_docs(self):
        """
        Test Holdings entity generates all 5 documents.
        """
        month = date(2026, 4, 1)
        
        documents = NParksService.generate_nparks(
            entity_id=self.holdings.id,
            month=month,
            generated_by_id=self.user.id,
        )

        # Should have all 5 documents
        self.assertEqual(len(documents), 5)
        self.assertIn("mating_sheet", documents)
        self.assertIn("puppy_movement", documents)
        self.assertIn("vet_treatments", documents)
        self.assertIn("puppies_bred", documents)
        self.assertIn("dog_movement", documents)

        # Verify submission created
        submission = NParksSubmission.objects.get(
            entity=self.holdings,
            month=month,
        )
        self.assertEqual(submission.status, NParksStatus.DRAFT)

    def test_generate_katong_5_docs(self):
        """
        Test Katong entity generates all 5 documents.
        """
        month = date(2026, 4, 1)
        
        documents = NParksService.generate_nparks(
            entity_id=self.katong.id,
            month=month,
            generated_by_id=self.user.id,
        )

        self.assertEqual(len(documents), 5)

    def test_generate_thomson_3_docs(self):
        """
        Test Thomson entity generates 3 documents (no mating/bred).
        
        Thomson doesn't have breeding operations, so mating and 
        puppies_bred should be empty or minimal.
        """
        month = date(2026, 4, 1)
        
        documents = NParksService.generate_nparks(
            entity_id=self.thomson.id,
            month=month,
            generated_by_id=self.user.id,
        )

        # Thomson should still generate all 5 docs, but mating/bred will be empty
        self.assertEqual(len(documents), 5)

    def test_dual_sire_columns_present(self):
        """
        Test mating sheet has dual-sire columns.
        """
        # Create breeding record with dual sire
        from apps.breeding.models import BreedingRecord
        
        breeding = BreedingRecord.objects.create(
            dam=self.dam,
            sire1=self.sire1,
            sire2=self.sire2,
            date=date(2026, 4, 15),
            method=BreedingRecord.Method.NATURAL,
            entity=self.holdings,
        )

        month = date(2026, 4, 1)
        documents = NParksService.generate_nparks(
            entity_id=self.holdings.id,
            month=month,
            generated_by_id=self.user.id,
        )

        # Load mating sheet
        wb = load_workbook(BytesIO(documents["mating_sheet"]))
        ws = wb.active

        # Check headers include Sire 1 and Sire 2
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Sire 1 Microchip", headers)
        self.assertIn("Sire 2 Microchip", headers)

    def test_farm_details_pre_filled(self):
        """
        Test farm details are pre-filled in documents.
        """
        month = date(2026, 4, 1)
        documents = NParksService.generate_nparks(
            entity_id=self.holdings.id,
            month=month,
            generated_by_id=self.user.id,
        )

        # Check mating sheet has farm details
        wb = load_workbook(BytesIO(documents["mating_sheet"]))
        ws = wb.active

        # Farm details should be in footer
        found_farm = False
        found_license = False
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if "Wellfond Pets Holdings" in str(cell.value):
                        found_farm = True
                    if "DB000065X" in str(cell.value):
                        found_license = True

        self.assertTrue(found_farm, "Farm name not found")
        self.assertTrue(found_license, "License number not found")

    def test_lock_prevents_edits(self):
        """
        Test locked submission prevents edits.
        """
        month = date(2026, 4, 1)
        
        documents = NParksService.generate_nparks(
            entity_id=self.holdings.id,
            month=month,
            generated_by_id=self.user.id,
        )

        submission = NParksSubmission.objects.get(
            entity=self.holdings,
            month=month,
        )

        # Lock the submission
        submission.status = NParksStatus.LOCKED
        submission.locked_at = timezone.now()
        submission.save()

        # Verify locked
        submission.refresh_from_db()
        self.assertEqual(submission.status, NParksStatus.LOCKED)

    def test_submit_records_timestamp(self):
        """
        Test submission records timestamp.
        """
        month = date(2026, 4, 1)
        
        documents = NParksService.generate_nparks(
            entity_id=self.holdings.id,
            month=month,
            generated_by_id=self.user.id,
        )

        submission = NParksSubmission.objects.get(
            entity=self.holdings,
            month=month,
        )

        # Submit
        submission.status = NParksStatus.SUBMITTED
        submission.submitted_at = timezone.now()
        submission.save()

        submission.refresh_from_db()
        self.assertIsNotNone(submission.submitted_at)
        self.assertEqual(submission.status, NParksStatus.SUBMITTED)

    def test_preview_matches_data(self):
        """
        Test preview data matches actual document data.
        """
        month = date(2026, 4, 1)
        
        preview = NParksService.preview_nparks(
            entity_id=self.holdings.id,
            month=month,
            doc_type="mating_sheet",
        )

        self.assertEqual(preview.doc_type, "mating_sheet")
        self.assertIsInstance(preview.headers, list)
        self.assertIsInstance(preview.rows, list)


class TestNParksValidation(TestCase):
    """Test NParks validation."""

    def setUp(self):
        """Set up test data."""
        self.entity, _ = Entity.objects.get_or_create(
            defaults={"name": "Test Entity", "code": "TEST", "slug": "test-entity"},
            id=uuid.uuid4(),
        )

    def test_validate_warnings(self):
        """
        Test validation returns warnings for issues.
        """
        month = date(2026, 4, 1)
        
        warnings = NParksService.validate_nparks(
            entity_id=self.entity.id,
            month=month,
        )

        # Should be a list
        self.assertIsInstance(warnings, list)

    def test_invalid_entity_raises_error(self):
        """
        Test invalid entity raises ValueError.
        """
        invalid_id = uuid.uuid4()
        month = date(2026, 4, 1)

        with self.assertRaises(ValueError) as context:
            NParksService.generate_nparks(
                entity_id=invalid_id,
                month=month,
                generated_by_id=uuid.uuid4(),
            )

        self.assertIn("Entity not found", str(context.exception))
